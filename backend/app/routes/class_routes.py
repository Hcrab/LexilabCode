from flask import Blueprint, request, jsonify, g, current_app
from bson.objectid import ObjectId
from werkzeug.security import generate_password_hash
from ..decorators import admin_required
from .student_routes import get_review_words
from .quiz_routes import compute_user_quiz_completion
import pytz
from datetime import datetime, timedelta
from bson.objectid import ObjectId

class_bp = Blueprint('class_bp', __name__)


def _get_class_if_teacher(class_id):
    """Fetch class doc and ensure current admin user is a teacher of it.
    Returns (doc, error_response) where error_response is a Flask response or None.
    """
    teacher_id = g.current_user.get('_id')
    try:
        class_object_id = ObjectId(class_id)
    except Exception:
        return None, (jsonify({'message': '无效的班级ID格式'}), 400)

    target_class = current_app.db.classes.find_one({'_id': class_object_id})
    if not target_class:
        return None, (jsonify({'message': '未找到指定班级'}), 404)

    # Only teachers who own this class can view/manage it
    teachers = target_class.get('teachers') or []
    if teacher_id not in teachers:
        return None, (jsonify({'message': '无权限访问该班级'}), 403)

    return target_class, None

@class_bp.route('/api/classes', methods=['GET'])
@admin_required
def get_teacher_classes():
    """
    Fetches all classes taught by the current logged-in teacher.
    """
    teacher_id = g.current_user.get('_id')
    try:
        classes = list(current_app.db.classes.find({'teachers': teacher_id}))
        
        for c in classes:
            c['_id'] = str(c['_id'])
            c['teachers'] = [str(tid) for tid in c['teachers']]
            c['students'] = [str(sid) for sid in c['students']]
            
        return jsonify(classes), 200
    except Exception as e:
        return jsonify({'message': '获取班级列表时发生错误', 'error': str(e)}), 500

@class_bp.route('/api/classes/<class_id>', methods=['GET'])
@admin_required
def get_class_details(class_id):
    """
    Fetches details for a single class, including a list of its students
    and a list of words that have already been assigned to the class.
    """
    try:
        # First, find the class and ensure ownership
        target_class, err = _get_class_if_teacher(class_id)
        if err:
            return err

        # Get the list of student IDs
        student_ids = target_class.get('students', [])

        # Find all students matching the IDs
        students_cursor = current_app.db.users.find(
            {'_id': {'$in': student_ids}},
            {'_id': 1, 'username': 1, 'nickname': 1, 'tier': 1, 'english_name': 1}  # Projection includes english_name
        )
        
        students_list = list(students_cursor)

        # Build assigned words set from class-level batches only
        batch_words = set()
        for b in target_class.get('assignment_word_batches', []) or []:
            for w in (b.get('words') or []):
                if isinstance(w, str) and w:
                    batch_words.add(w)
        assigned_words = sorted(list(batch_words))

        # Prepare the response
        target_class['students'] = students_list
        target_class['assigned_words'] = sorted(list(assigned_words))
        
        # Convert all ObjectIds to strings for JSON
        target_class['_id'] = str(target_class['_id'])
        target_class['teachers'] = [str(tid) for tid in target_class['teachers']]
        for student in target_class['students']:
            student['_id'] = str(student['_id'])

        # Remove raw history fields from the final response (not used now)
        if 'assigned_words_history' in target_class:
            del target_class['assigned_words_history']
        if 'assignment_word_batches' in target_class:
            del target_class['assignment_word_batches']

        return jsonify(target_class), 200
    except Exception as e:
        return jsonify({'message': '获取班级详情时发生错误', 'error': str(e)}), 500

@class_bp.route('/api/classes/<class_id>/words', methods=['GET'])
@admin_required
def get_class_words(class_id):
    """
    Fetches a unique list of all words (mastered and to-be-mastered)
    for all students in a given class.
    """
    try:
        target_class, err = _get_class_if_teacher(class_id)
        if err:
            return err

        student_ids = target_class.get('students', [])
        if not student_ids:
            return jsonify([]), 200 # Return empty list if no students

        # Fetch all relevant students in one query
        students_cursor = current_app.db.users.find(
            {'_id': {'$in': student_ids}},
            {'to_be_mastered.word': 1, 'words_mastered.word': 1, '_id': 0} # Projection
        )
        
        all_words = set()
        for student in students_cursor:
            for item in student.get('to_be_mastered', []):
                all_words.add(item['word'])
            for item in student.get('words_mastered', []):
                all_words.add(item['word'])
        
        return jsonify(sorted(list(all_words))), 200

    except Exception as e:
        return jsonify({'message': '获取班级词汇列表时发生错误', 'error': str(e)}), 500

@class_bp.route('/api/classes', methods=['POST'])
@admin_required
def create_class():
    """
    Creates a new class.
    The teacher creating the class is automatically added to the 'teachers' list.
    """
    data = request.get_json()
    if not data or not data.get('name'):
        return jsonify({'message': '请求中缺少班级名称'}), 400

    teacher_id = g.current_user.get('_id')

    new_class = {
        "name": data['name'],
        "teachers": [teacher_id],
        "students": []
    }

    try:
        result = current_app.db.classes.insert_one(new_class)
        created_class = current_app.db.classes.find_one({'_id': result.inserted_id})
        
        # Convert ObjectId to string for the response
        created_class['_id'] = str(created_class['_id'])
        created_class['teachers'] = [str(tid) for tid in created_class['teachers']]
        
        return jsonify(created_class), 201
    except Exception as e:
        return jsonify({'message': '创建班级时发生错误', 'error': str(e)}), 500

@class_bp.route('/api/classes/join', methods=['POST'])
@admin_required
def join_class_as_teacher():
    """
    Adds current teacher to an existing class's teachers list by class_id.
    """
    data = request.get_json(silent=True) or {}
    class_id = data.get('class_id')
    if not class_id:
        return jsonify({'message': '缺少class_id'}), 400
    try:
        cid = ObjectId(class_id)
    except Exception:
        return jsonify({'message': '无效的班级ID格式'}), 400
    teacher_id = g.current_user.get('_id')
    # Add teacher to class if exists
    res = current_app.db.classes.update_one({'_id': cid}, {'$addToSet': {'teachers': teacher_id}})
    if res.matched_count == 0:
        return jsonify({'message': '未找到指定班级'}), 404
    return jsonify({'message': '已加入班级'}), 200

@class_bp.route('/api/classes/<class_id>/students', methods=['POST'])
@admin_required
def add_students_to_class(class_id):
    """
    Adds a list of students to a specific class.
    """
    data = request.get_json()
    student_ids = data.get('student_ids')

    if not student_ids or not isinstance(student_ids, list):
        return jsonify({'message': '请求中缺少学生ID列表'}), 400

    try:
        target_class, err = _get_class_if_teacher(class_id)
        if err:
            return err
        student_object_ids = [ObjectId(sid) for sid in student_ids]
        class_object_id = target_class['_id']
    except Exception:
        return jsonify({'message': '无效的班级ID或学生ID格式'}), 400

    # Add students to the class's student list, ensuring no duplicates
    result = current_app.db.classes.update_one(
        {'_id': class_object_id},
        {'$addToSet': {'students': {'$each': student_object_ids}}}
    )

    if result.matched_count == 0:
        return jsonify({'message': '未找到指定班级'}), 404

    return jsonify({'message': f'成功将 {result.modified_count} 名新学生添加到班级'}), 200


@class_bp.route('/api/classes/<class_id>/students/<student_id>', methods=['DELETE'])
@admin_required
def remove_student_from_class(class_id, student_id):
    """
    Removes a student from a specific class.
    """
    try:
        target_class, err = _get_class_if_teacher(class_id)
        if err:
            return err
        class_object_id = target_class['_id']
        student_object_id = ObjectId(student_id)
    except Exception:
        return jsonify({'message': '无效的班级ID或学生ID格式'}), 400

    result = current_app.db.classes.update_one(
        {'_id': class_object_id},
        {'$pull': {'students': student_object_id}}
    )

    if result.matched_count == 0:
        return jsonify({'message': '未找到指定班级'}), 404
    
    if result.modified_count == 0:
        return jsonify({'message': '该学生不在该班级中'}), 404

    return jsonify({'message': '学生已成功从班级中移除'}), 200


@class_bp.route('/api/classes/<class_id>/assigned-vocab', methods=['GET'])
@admin_required
def get_class_assigned_vocab(class_id):
    """
    For a given class, returns a representative student's current vocabulary:
    - to_be_mastered (list of words)
    - words_mastered (list of words)
    - combined (unique union)
    Assumes all students in the class are assigned the same words.
    """
    target_class, err = _get_class_if_teacher(class_id)
    if err:
        return err

    student_ids = target_class.get('students', [])
    if not student_ids:
        return jsonify({'to_be_mastered': [], 'words_mastered': [], 'combined': []}), 200

    # Pick the first student as representative
    representative_id = student_ids[0]
    student = current_app.db.users.find_one(
        {'_id': representative_id},
        {'to_be_mastered.word': 1, 'words_mastered.word': 1}
    )

    if not student:
        return jsonify({'to_be_mastered': [], 'words_mastered': [], 'combined': []}), 200

    tbm = [e.get('word') for e in student.get('to_be_mastered', []) if e.get('word')]
    mastered = [e.get('word') for e in student.get('words_mastered', []) if e.get('word')]
    combined = sorted(list(set(tbm + mastered)))

    return jsonify({
        'to_be_mastered': sorted(list(set(tbm))),
        'words_mastered': sorted(list(set(mastered))),
        'combined': combined
    }), 200

@class_bp.route('/api/classes/<class_id>/assignment-history', methods=['GET'])
@admin_required
def get_class_assignment_history(class_id):
    """
    Returns the class-level word assignment history grouped by assigned_date.
    Uses a representative student's vocabulary lists to infer the class history
    (since assignments are applied to all students simultaneously).
    Response format: [{ 'date': 'YYYY-MM-DD', 'words': ['word1','word2', ...] }]
    """
    target_class, err = _get_class_if_teacher(class_id)
    if err:
        return err

    # Preferred: read from class batch records if present
    batches = target_class.get('assignment_word_batches', []) or []
    # Group by date and merge words
    grouped = {}
    for b in batches:
        date = b.get('assigned_date')
        words = b.get('words') or []
        if not date or not isinstance(words, list):
            continue
        grouped.setdefault(date, set()).update([w for w in words if isinstance(w, str) and w])
    result = [
        {'date': d, 'words': sorted(list(ws))}
        for d, ws in grouped.items()
    ]
    result.sort(key=lambda x: x['date'], reverse=True)
    return jsonify(result), 200


@class_bp.route('/api/classes/<class_id>/bulk-import-students', methods=['POST'])
@admin_required
def bulk_import_students(class_id):
    """
    Bulk import students into a class.
    Supports two modes:
      1) JSON body: { prefix: string, count: number } to auto-generate usernames
      2) Multipart upload: file=CSV/XLSX, with headers: username,password,e_name
         - username: required, unique key
         - password: optional; default '123456' if empty
         - e_name: optional; stored as english_name
    """
    # Resolve and authorize class
    try:
        target_class, err = _get_class_if_teacher(class_id)
        if err:
            return err
        class_object_id = target_class['_id']
    except Exception:
        return jsonify({'message': '无效的班级ID格式'}), 400

    created_students = []
    existing_students = []
    invalid_rows = []
    new_student_ids = []

    ct = (request.content_type or '').lower()
    # Multipart form with file upload
    if 'multipart/form-data' in ct:
        up = request.files.get('file')
        if not up or not getattr(up, 'filename', ''):
            return jsonify({'message': '缺少上传文件(file)'}), 400
        filename = up.filename or ''
        name_lower = filename.lower()

        rows = []
        try:
            if name_lower.endswith('.csv'):
                import io, csv
                content = up.stream.read()
                try:
                    text = content.decode('utf-8')
                except Exception:
                    # Fallback to gbk commonly used in CN Excel exports
                    text = content.decode('gbk', errors='ignore')
                reader = csv.DictReader(io.StringIO(text))
                for idx, r in enumerate(reader, start=2):  # header at line 1
                    rows.append((idx, r))
            elif name_lower.endswith('.xlsx'):
                try:
                    import openpyxl  # optional dependency
                except Exception:
                    return jsonify({'message': '服务器未安装xlsx解析库，请改用CSV或联系管理员安装openpyxl'}), 400
                wb = openpyxl.load_workbook(up.stream, data_only=True)
                ws = wb.active
                headers = []
                for j, cell in enumerate(ws[1], start=1):
                    headers.append(str(cell.value or '').strip())
                header_map = {h.lower(): i for i, h in enumerate(headers)}
                for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                    rdict = {}
                    for key in ('username', 'password', 'e_name'):
                        if key in header_map:
                            rdict[key] = row[header_map[key]]
                        else:
                            rdict[key] = None
                    rows.append((i, rdict))
            else:
                return jsonify({'message': '仅支持CSV或XLSX文件'}), 400
        except Exception as e:
            return jsonify({'message': f'解析文件失败: {str(e)}'}), 400

        # Process rows
        MAX_ROWS = 500
        if len(rows) > MAX_ROWS:
            return jsonify({'message': f'单次最多导入{MAX_ROWS}行'}), 400

        for line_no, r in rows:
            username = str(r.get('username') or '').strip()
            password_raw = str(r.get('password') or '').strip()
            e_name = str(r.get('e_name') or '').strip() if r.get('e_name') is not None else ''
            if not username:
                invalid_rows.append({'line': line_no, 'error': 'username缺失'})
                continue
            if current_app.db.users.find_one({'username': username}):
                existing_students.append(username)
                continue
            try:
                doc = {
                    'username': username,
                    'password': generate_password_hash(password_raw or '123456'),
                    'role': 'user',
                    'tier': 'tier_3',
                    'ai_calls': 0,
                    'words_mastered': [],
                    'to_be_mastered': []
                }
                if e_name:
                    doc['english_name'] = e_name
                ins = current_app.db.users.insert_one(doc)
                created_students.append(username)
                new_student_ids.append(ins.inserted_id)
            except Exception as e:
                invalid_rows.append({'line': line_no, 'error': f'创建失败: {str(e)}'})

    else:
        # Legacy JSON mode: prefix + count
        data = request.get_json(silent=True) or {}
        prefix = data.get('prefix')
        count = data.get('count')
        if not prefix or count is None:
            return jsonify({'message': '请求中缺少前缀或数量'}), 400
        try:
            count = int(count)
            if count <= 0 or count > 200:
                return jsonify({'message': '数量必须在1到200之间'}), 400
        except (ValueError, TypeError):
            return jsonify({'message': '数量必须是一个有效的整数'}), 400

        for i in range(1, count + 1):
            username = f"{prefix}{i:02d}"
            if current_app.db.users.find_one({'username': username}):
                existing_students.append(username)
                continue
            new_student = {
                'username': username,
                'password': generate_password_hash('123456'),
                'role': 'user',
                'tier': 'tier_3',
                'ai_calls': 0,
                'words_mastered': [],
                'to_be_mastered': []
            }
            res = current_app.db.users.insert_one(new_student)
            created_students.append(username)
            new_student_ids.append(res.inserted_id)

    # Post-create: track default wordbooks and add to class
    if new_student_ids:
        try:
            regex = {'$regex': '(A1|A2|B1|B2|C1)', '$options': 'i'}
            cur = current_app.db.wordbooks.find({
                'title': regex,
                '$or': [
                    {'accessibility': 'public'},
                    {'accessibility': {'$exists': False}}
                ]
            }, {'_id': 1})
            track_ids = [doc['_id'] for doc in cur]
            current_app.db.users.update_many(
                {'_id': {'$in': new_student_ids}},
                {'$set': {'tracked_wordbooks': track_ids}}
            )
        except Exception:
            pass
        current_app.db.classes.update_one(
            {'_id': class_object_id},
            {'$addToSet': {'students': {'$each': new_student_ids}}}
        )

    return jsonify({
        'message': '批量导入完成',
        'created_students': created_students,
        'existing_students': existing_students,
        'invalid_rows': invalid_rows
    }), 201

@class_bp.route('/api/classes/<class_id>/assign-words', methods=['POST'])
@admin_required
def assign_words_to_class(class_id):
    """
    Assigns a list of words to all students in a class and records
    the words in the class's assignment history.
    """
    data = request.get_json()
    words_to_assign = data.get('words')

    if not words_to_assign or not isinstance(words_to_assign, list):
        return jsonify({'message': '请求中缺少单词列表'}), 400

    # Get class and ensure ownership
    target_class, err = _get_class_if_teacher(class_id)
    if err:
        return err
    
    student_ids = target_class.get('students', [])
    if not student_ids:
        return jsonify({'message': '该班级没有学生'}), 200

    # Prepare the assignment entries with correct timezone handling
    beijing_tz = pytz.timezone('Asia/Shanghai')
    now_in_beijing = datetime.now(beijing_tz)
    
    assigned_date = now_in_beijing.strftime('%Y-%m-%d')
    # The due date is the end of the *next* day in Beijing time
    due_date = (now_in_beijing + timedelta(days=1)).strftime('%Y-%m-%d')

    new_assignments = [
        {'word': word, 'assigned_date': assigned_date, 'due_date': due_date, 'source': 'teacher'}
        for word in words_to_assign
    ]

    # Update all students in the class
    result = current_app.db.users.update_many(
        {'_id': {'$in': student_ids}},
        {'$addToSet': {'to_be_mastered': {'$each': new_assignments}}}
    )

    # Record vocab_mission for each student (unique by word)
    try:
        students = list(current_app.db.users.find({'_id': {'$in': student_ids}}, {'_id': 1, 'vocab_mission.word': 1}))
        for s in students:
            existed = set()
            for m in (s.get('vocab_mission') or []):
                w = m.get('word') if isinstance(m, dict) else None
                if w:
                    existed.add(w)
            vm_new = [{'word': w, 'assigned_date': assigned_date, 'source': 'teacher'} for w in words_to_assign if isinstance(w, str) and w and w not in existed]
            if vm_new:
                current_app.db.users.update_one({'_id': s['_id']}, {'$push': {'vocab_mission': {'$each': vm_new}}})
    except Exception:
        pass

    # Append a batch record for precise per-class history by date
    current_app.db.classes.update_one(
        {'_id': target_class['_id']},
        {'$push': {
            'assignment_word_batches': {
                'assigned_date': assigned_date,
                'words': words_to_assign
            }
        }}
    )

    return jsonify({
        'message': f'成功为 {result.modified_count} 名学生布置了 {len(words_to_assign)} 个单词。'
    }), 200


@class_bp.route('/api/classes/<class_id>/learning-goal', methods=['PUT'])
@admin_required
def set_class_learning_goal(class_id):
    """
    Set a daily learning goal for all students in the class and lock it so students cannot change it.
    Body: { goal: int }
    """
    data = request.get_json(silent=True) or {}
    try:
        goal = int(data.get('goal'))
    except Exception:
        return jsonify({'message': 'Missing or invalid learning goal'}), 400
    if goal < 0 or goal > 500:
        return jsonify({'message': 'Learning goal must be between 0 and 500'}), 400

    # Ensure class exists and current teacher owns it
    target_class, err = _get_class_if_teacher(class_id)
    if err:
        return err
    student_ids = target_class.get('students', []) or []
    if not student_ids:
        return jsonify({'message': '班级暂无学生'}), 200

    # Apply learning_goal and lock on all students
    res = current_app.db.users.update_many(
        {'_id': {'$in': student_ids}},
        {'$set': {
            'learning_goal': goal,
            'learning_goal_locked': True,
            'learning_goal_locked_by_class': target_class['_id']
        }}
    )
    return jsonify({'message': f'Applied and locked learning goal for {res.modified_count} students', 'applied': int(res.modified_count), 'goal': goal}), 200


@class_bp.route('/api/classes/<class_id>/learning-goal', methods=['GET'])
@admin_required
def get_class_learning_goal(class_id):
    """
    Return the current class learning goal snapshot inferred from students.
    If a majority (or any) students are locked by this class, returns the most common goal among them.
    Response: { goal: number|null, locked_count: number, total: number }
    """
    target_class, err = _get_class_if_teacher(class_id)
    if err:
        return err
    student_ids = target_class.get('students', []) or []
    if not student_ids:
        return jsonify({'goal': None, 'locked_count': 0, 'total': 0}), 200

    # Fetch student goal + lock info
    cur = current_app.db.users.find({'_id': {'$in': student_ids}}, {
        'learning_goal': 1,
        'learning_goal_locked': 1,
        'learning_goal_locked_by_class': 1
    })
    goals = []
    locked_goals = []
    for u in cur:
        g = None
        try:
            g = int(u.get('learning_goal') or 0)
        except Exception:
            g = 0
        goals.append(g)
        if u.get('learning_goal_locked') and u.get('learning_goal_locked_by_class') == target_class['_id']:
            locked_goals.append(g)

    locked_count = len(locked_goals)
    total = len(goals)

    # Prefer the most common locked goal when present; else fall back to overall most common goal
    def _most_common(arr):
        if not arr:
            return None
        freq = {}
        for x in arr:
            freq[x] = freq.get(x, 0) + 1
        # return value with highest frequency
        return max(freq.items(), key=lambda kv: kv[1])[0]

    goal = _most_common(locked_goals) if locked_goals else _most_common(goals)
    return jsonify({'goal': goal, 'locked_count': locked_count, 'total': total}), 200


@class_bp.route('/api/classes/<class_id>/students/<student_id>/assign-words', methods=['POST'])
@admin_required
def assign_words_to_student(class_id, student_id):
    data = request.get_json() or {}
    words = data.get('words') or []
    if not isinstance(words, list) or not words:
        return jsonify({'message': '请求中缺少单词列表'}), 400
    try:
        target_class, err = _get_class_if_teacher(class_id)
        if err:
            return err
        class_oid = target_class['_id']
        student_oid = ObjectId(student_id)
    except Exception:
        return jsonify({'message': '无效的班级ID或学生ID'}), 400
    # Verify student belongs to class
    cls = current_app.db.classes.find_one({'_id': class_oid, 'students': student_oid}, {'_id': 1})
    if not cls:
        return jsonify({'message': '学生不属于该班级'}), 404
    beijing_tz = pytz.timezone('Asia/Shanghai')
    now_in_beijing = datetime.now(beijing_tz)
    assigned_date = now_in_beijing.strftime('%Y-%m-%d')
    due_date = (now_in_beijing + timedelta(days=1)).strftime('%Y-%m-%d')
    # Exclude duplicates from student's existing lists
    user_doc = current_app.db.users.find_one({'_id': student_oid}, {'to_be_mastered.word': 1, 'words_mastered.word': 1, 'vocab_mission.word': 1}) or {}
    existing = set()
    for e in (user_doc.get('to_be_mastered') or []):
        w = e.get('word') if isinstance(e, dict) else e
        if isinstance(w, str): existing.add(w)
    for e in (user_doc.get('words_mastered') or []):
        w = e.get('word') if isinstance(e, dict) else e
        if isinstance(w, str): existing.add(w)
    to_add = [w for w in words if isinstance(w, str) and w and w not in existing]
    entries = [{'word': w, 'assigned_date': assigned_date, 'due_date': due_date, 'source': 'teacher'} for w in to_add]
    if entries:
        current_app.db.users.update_one({'_id': student_oid}, {'$addToSet': {'to_be_mastered': {'$each': entries}}})
    # Update vocab_mission unique by word
    existed_vm = set(m.get('word') for m in (user_doc.get('vocab_mission') or []) if isinstance(m, dict) and m.get('word'))
    vm_to_add = [{'word': w, 'assigned_date': assigned_date, 'source': 'teacher'} for w in words if isinstance(w, str) and w and w not in existed_vm]
    if vm_to_add:
        current_app.db.users.update_one({'_id': student_oid}, {'$push': {'vocab_mission': {'$each': vm_to_add}}})
    return jsonify({'message': f'已为学生布置 {len(entries)} 个新单词', 'added': len(entries), 'requested': len(words)}), 200


@class_bp.route('/api/classes/<class_id>/students/<student_id>/assign-from-wordbook', methods=['POST'])
@admin_required
def assign_from_wordbook_to_student(class_id, student_id):
    data = request.get_json() or {}
    wordbook_id = data.get('wordbook_id')
    count = int(data.get('count', 10))
    if not wordbook_id:
        return jsonify({'message': '缺少词库ID'}), 400
    try:
        class_oid = ObjectId(class_id)
        student_oid = ObjectId(student_id)
        wb_oid = ObjectId(wordbook_id)
    except Exception:
        return jsonify({'message': '无效的ID'}), 400
    cls = current_app.db.classes.find_one({'_id': class_oid, 'students': student_oid}, {'_id': 1})
    if not cls:
        return jsonify({'message': '学生不属于该班级'}), 404
    wb = current_app.db.wordbooks.find_one({'_id': wb_oid}, {'entries': 1, 'title': 1})
    if not wb:
        return jsonify({'message': '词库不存在'}), 404
    # Build exclusion and valid sets
    user_doc = current_app.db.users.find_one({'_id': student_oid}, {'to_be_mastered.word': 1, 'words_mastered.word': 1, 'vocab_mission.word': 1}) or {}
    existing = set()
    for e in (user_doc.get('to_be_mastered') or []):
        w = e.get('word') if isinstance(e, dict) else e
        if isinstance(w, str): existing.add(w)
    for e in (user_doc.get('words_mastered') or []):
        w = e.get('word') if isinstance(e, dict) else e
        if isinstance(w, str): existing.add(w)
    valid_words = set(doc.get('word') for doc in current_app.db.words.find({}, {'word': 1}) if doc.get('word'))
    entries = wb.get('entries') or []
    entries_sorted = sorted(entries, key=lambda e: e.get('number', 0))
    candidates = []
    for e in entries_sorted:
        w = e.get('word') if isinstance(e, dict) else None
        if not w: continue
        if w in existing: continue
        if w not in valid_words: continue
        if w in candidates: continue
        candidates.append(w)
        if len(candidates) >= count: break
    if not candidates:
        return jsonify({'message': '没有可添加的新单词', 'added': 0, 'words': []}), 200
    # Assign
    beijing_tz = pytz.timezone('Asia/Shanghai')
    now_in_beijing = datetime.now(beijing_tz)
    assigned_date = now_in_beijing.strftime('%Y-%m-%d')
    due_date = (now_in_beijing + timedelta(days=1)).strftime('%Y-%m-%d')
    tbm_entries = [{'word': w, 'assigned_date': assigned_date, 'due_date': due_date, 'source': 'teacher'} for w in candidates]
    current_app.db.users.update_one({'_id': student_oid}, {'$addToSet': {'to_be_mastered': {'$each': tbm_entries}}})
    # vocab_mission unique by word
    existed_vm = set(m.get('word') for m in (user_doc.get('vocab_mission') or []) if isinstance(m, dict) and m.get('word'))
    vm_new = [{'word': w, 'assigned_date': assigned_date, 'source': 'teacher'} for w in candidates if w not in existed_vm]
    if vm_new:
        current_app.db.users.update_one({'_id': student_oid}, {'$push': {'vocab_mission': {'$each': vm_new}}})
    return jsonify({'message': f'成功加入 {len(tbm_entries)} 个单词', 'added': len(tbm_entries), 'words': candidates}), 200


@class_bp.route('/api/classes/<class_id>/students/<student_id>/vocab-mission-history', methods=['GET'])
@admin_required
def get_student_vocab_mission_history(class_id, student_id):
    try:
        target_class, err = _get_class_if_teacher(class_id)
        if err:
            return err
        class_oid = target_class['_id']
        student_oid = ObjectId(student_id)
    except Exception:
        return jsonify({'message': '无效的ID'}), 400
    cls = current_app.db.classes.find_one({'_id': class_oid, 'students': student_oid}, {'_id': 1})
    if not cls:
        return jsonify({'message': '学生不属于该班级'}), 404
    doc = current_app.db.users.find_one({'_id': student_oid}, {'vocab_mission': 1, 'username': 1, 'nickname': 1}) or {}
    missions = doc.get('vocab_mission') or []
    # sort desc by assigned_date
    try:
        missions.sort(key=lambda x: x.get('assigned_date', ''), reverse=True)
    except Exception:
        pass
    return jsonify({'student': {'_id': str(student_oid), 'username': doc.get('username'), 'nickname': doc.get('nickname')}, 'vocab_mission': missions}), 200

@class_bp.route('/api/classes/<class_id>/stats', methods=['GET'])
@admin_required
def get_class_stats(class_id):
    """
    Calculates and returns comprehensive statistics for all students in a class.
    The calculation is based on a date range, which defaults to the last 7 days.
    - Real-time learning/review completion status for the day.
    - Historical learning/review completion rates within the date range.
    - Assignment completion rates.
    """
    try:
        # --- Date Range Setup ---
        start_date_str = request.args.get('start_date')
        beijing_tz = pytz.timezone('Asia/Shanghai')

        # 1. Get the class first (used to determine default start date from first assignment)
        target_class, err = _get_class_if_teacher(class_id)
        if err:
            return err

        # Determine default start date: earliest assigned_date from class-level batches
        default_start = None
        try:
            batches = target_class.get('assignment_word_batches', []) or []
            dates = [b.get('assigned_date') for b in batches if isinstance(b, dict) and b.get('assigned_date')]
            if dates:
                earliest = min(dates)
                default_start = datetime.strptime(earliest, '%Y-%m-%d').astimezone(beijing_tz)
        except Exception:
            default_start = None

        if start_date_str:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').astimezone(beijing_tz)
        elif default_start:
            start_date = default_start
        else:
            # Fallback to the last 7 days
            start_date = datetime.now(beijing_tz) - timedelta(days=6)

        end_date = datetime.now(beijing_tz)
        # Ensure start_date is at the beginning of the day
        start_date = start_date.replace(hour=0, minute=0, second=0, microsecond=0)

        total_days_in_range = (end_date.date() - start_date.date()).days + 1
        if total_days_in_range <= 0:
            total_days_in_range = 1

        # 2. Get the class's student IDs
        student_ids = target_class.get('students', [])
        if not student_ids:
            return jsonify([]), 200

        # 3. Get all student documents at once for efficiency
        students_cursor = current_app.db.users.find({'_id': {'$in': student_ids}})
        students_map = {s['_id']: s for s in students_cursor}

        # Assignment completion now unified to quiz completion; no legacy submissions lookup

        # 6. Combine all data for the final response
        final_stats = []
        start_date_str_for_compare = start_date.strftime('%Y-%m-%d')

        for student_id, student_doc in students_map.items():
            # Real-time stats
            completed_today_learning = len(student_doc.get('to_be_mastered', [])) == 0
            review_words = get_review_words(student_doc)
            completed_today_review = len(review_words) == 0

            # --- Lazy, in-app daily completion tracking (no crontab) ---
            # If student has cleared today's tasks, persist today's date into the arrays.
            today_str = datetime.now(beijing_tz).strftime('%Y-%m-%d')
            updated_learn_days = set(student_doc.get('complete_exercise_day', []) or [])
            updated_review_days = set(student_doc.get('complete_revision_day', []) or [])

            if completed_today_learning and today_str not in updated_learn_days:
                try:
                    current_app.db.users.update_one({'_id': student_id}, {'$addToSet': {'complete_exercise_day': today_str}})
                    updated_learn_days.add(today_str)
                except Exception as _:
                    pass
            if completed_today_review and today_str not in updated_review_days:
                try:
                    current_app.db.users.update_one({'_id': student_id}, {'$addToSet': {'complete_revision_day': today_str}})
                    updated_review_days.add(today_str)
                except Exception as _:
                    pass

            # Historical stats derived from recorded days (including today's lazy add)
            learning_days_completed = [d for d in updated_learn_days if d >= start_date_str_for_compare]
            review_days_completed = [d for d in updated_review_days if d >= start_date_str_for_compare]

            learning_rate = (len(learning_days_completed) / total_days_in_range * 100)
            review_rate = (len(review_days_completed) / total_days_in_range * 100)

            # Quiz-based completion stats (unified)
            try:
                comp = compute_user_quiz_completion(student_doc.get('username'))
                assignment_rate = comp.get('completion_rate', 0)
            except Exception:
                assignment_rate = 0

            final_stats.append({
                'student_id': str(student_id),
                'username': student_doc['username'],
                'nickname': student_doc.get('nickname', ''),
                'completed_today_learning': completed_today_learning,
                'completed_today_review': completed_today_review,
                'learning_completion_rate': round(learning_rate, 2),
                'review_completion_rate': round(review_rate, 2),
                'assignment_completion_rate': round(assignment_rate, 2)
            })
            
        return jsonify(final_stats), 200

    except Exception as e:
        current_app.logger.error(f"获取班级统计数据时出错: {e}")
        return jsonify({'message': '获取班级统计数据时发生内部错误', 'error': str(e)}), 500

@class_bp.route('/api/classes/<class_id>/exams', methods=['GET'])
@admin_required
def get_class_exams(class_id):
    """
    Fetches all exams for a given class and calculates their completion rates and average scores.
    """
    try:
        # 1. Get the class to find the number of students and ensure ownership
        target_class, err = _get_class_if_teacher(class_id)
        if err:
            return err
        
        student_ids = target_class.get('students', [])
        num_students = len(student_ids)
        if num_students == 0:
            return jsonify([]), 200

        # 2. Find all exams published for this class
        exams = list(current_app.db.assignments.find({
            'class_id': target_class['_id'],
            'status': 'published'
        }))

        if not exams:
            return jsonify([]), 200

        exam_ids = [e['_id'] for e in exams]

        # 3. Aggregation to count unique submissions (completion rate)
        completion_pipeline = [
            {'$match': {'assignment_id': {'$in': exam_ids}}},
            {'$group': {'_id': {'assignment_id': '$assignment_id', 'student_id': '$student_id'}}},
            {'$group': {'_id': '$_id.assignment_id', 'submission_count': {'$sum': 1}}}
        ]
        submission_counts = list(current_app.db.submissions.aggregate(completion_pipeline))
        submission_map = {item['_id']: item['submission_count'] for item in submission_counts}

        # 4. Aggregation to calculate average score from first submissions
        avg_score_pipeline = [
            {'$match': {'assignment_id': {'$in': exam_ids}, 'student_id': {'$in': student_ids}}},
            {'$sort': {'submitted_at': 1}},
            {'$group': {
                '_id': {'assignment_id': '$assignment_id', 'student_id': '$student_id'},
                'first_score': {'$first': '$total_score'}
            }},
            {'$group': {
                '_id': '$_id.assignment_id',
                'average_score': {'$avg': '$first_score'}
            }}
        ]
        avg_scores = list(current_app.db.submissions.aggregate(avg_score_pipeline))
        avg_score_map = {item['_id']: item['average_score'] for item in avg_scores}

        # 5. Format the response
        exam_stats = []
        for exam in exams:
            exam_id = exam['_id']
            submissions = submission_map.get(exam_id, 0)
            completion_rate = (submissions / num_students * 100) if num_students > 0 else 0
            average_score = avg_score_map.get(exam_id) # This can be None
            
            exam_stats.append({
                '_id': str(exam_id),
                'name': exam['name'],
                'created_at': exam['created_at'],
                'completion_rate': round(completion_rate, 2),
                'average_score': round(average_score, 2) if average_score is not None else 0
            })

        return jsonify(exam_stats), 200
    except Exception as e:
        current_app.logger.error(f"获取班级测验列表时出错: {e}")
        return jsonify({'message': '获取班级测验列表时发生内部错误', 'error': str(e)}), 500

@class_bp.route('/api/classes/<class_id>/secret-wordbook-from-box', methods=['POST'])
@admin_required
def set_class_secret_from_box(class_id):
    """
    Apply a teacher's secret box wordbook to all students in the class.
    For each student: ensure a private secret wordbook exists, rename to box title,
    merge entries, and set it as the only tracked wordbook.
    """
    target_class, err = _get_class_if_teacher(class_id)
    if err:
        return err
    data = request.get_json(silent=True) or {}
    teacher_id = g.current_user.get('_id')
    # 支持多个词库：优先使用 box_ids（数组）；否则回退到单个 box_id
    box_ids = data.get('box_ids')
    boxes = []
    if isinstance(box_ids, list) and len(box_ids) > 0:
        try:
            bids = [ObjectId(x) for x in box_ids]
        except Exception:
            return jsonify({'message': '无效词库ID'}), 400
        boxes = list(current_app.db.wordbooks.find({'_id': {'$in': bids}, 'creator_id': teacher_id, 'accessibility': 'teacher_secret'}))
        if not boxes or len(boxes) != len(bids):
            return jsonify({'message': '词库不存在或无权限'}), 404
    else:
        box_id = data.get('box_id')
        try:
            bid = ObjectId(box_id)
        except Exception:
            return jsonify({'message': '无效词库ID'}), 400
        box = current_app.db.wordbooks.find_one({'_id': bid, 'creator_id': teacher_id, 'accessibility': 'teacher_secret'})
        if not box:
            return jsonify({'message': '词库不存在或无权限'}), 404
        boxes = [box]

    student_ids = target_class.get('students') or []
    if not student_ids:
        return jsonify({'message': '班级暂无学生'}), 200

    applied = 0
    for sid in student_ids:
        stu = current_app.db.users.find_one({'_id': sid, 'role': 'user'})
        if not stu:
            continue
        wb = current_app.db.wordbooks.find_one({'creator_id': sid, 'accessibility':'private', 'title': '秘制词库'})
        if not wb:
            r = current_app.db.wordbooks.insert_one({
                'title': '秘制词库',
                'description': f"由老师为 {stu.get('username','student')} 指定的秘制词库",
                'categories': [],
                'entries': [],
                'creator_id': sid,
                'accessibility': 'private'
            })
            wb = current_app.db.wordbooks.find_one({'_id': r.inserted_id})
        wb_id = wb.get('_id')

        # Rename to first selected box title（若有），避免频繁改名
        try:
            first_title = (boxes[0] or {}).get('title', '秘制词库') if boxes else '秘制词库'
            current_app.db.wordbooks.update_one({'_id': wb_id}, {'$set': {'title': first_title}})
        except Exception:
            pass

        # Merge entries from all selected boxes into student's secret book
        wb_doc = current_app.db.wordbooks.find_one({'_id': wb_id}) or {}
        entries = wb_doc.get('entries') or []
        already = set(e.get('word') for e in entries if isinstance(e, dict))
        max_number = 0
        for e in entries:
            try:
                n = int(e.get('number') or 0)
                if n > max_number:
                    max_number = n
            except Exception:
                continue

        new_entries = []
        for bx in boxes:
            box_entries = bx.get('entries') or []
            for e in box_entries:
                try:
                    w = e.get('word')
                    if isinstance(w, str) and w and w not in already:
                        max_number += 1
                        new_entries.append({'number': max_number, 'word': w, 'tags': []})
                        already.add(w)
                except Exception:
                    continue
        if new_entries:
            current_app.db.wordbooks.update_one({'_id': wb_id}, {'$push': {'entries': {'$each': new_entries}}})

        # Track this wordbook (retain student's other follows) and lock by teacher
        current_app.db.users.update_one({'_id': sid}, {'$addToSet': {'tracked_wordbooks': wb_id}})
        try:
            current_app.db.wordbooks.update_one({'_id': wb_id}, {'$set': {'locked_by_teacher': True}})
        except Exception:
            pass
        applied += 1

    titles = [b.get('title') for b in boxes if isinstance(b, dict) and b.get('title')]
    return jsonify({'message': f'已为 {applied} 名学生设置秘制词库：' + ', '.join(titles), 'applied': applied, 'box_ids': [str(b.get('_id')) for b in boxes]}), 200

@class_bp.route('/api/classes/<class_id>/students/<student_id>/exam-history', methods=['GET'])
@admin_required
def get_student_exam_history(class_id, student_id):
    """
    Fetches the exam submission history for a specific student in a specific class.
    """
    try:
        class_object_id = ObjectId(class_id)
        student_object_id = ObjectId(student_id)
    except Exception:
        return jsonify({'message': '无效的班级ID或学生ID格式'}), 400

    try:
        # 1. Find all submissions by the student for the given class
        submissions = list(current_app.db.submissions.find({
            'student_id': student_object_id,
            'class_id': class_object_id
        }).sort('submitted_at', -1)) # Sort by most recent first

        if not submissions:
            return jsonify([]), 200

        # 2. Get the corresponding assignment (exam) details
        assignment_ids = [s['assignment_id'] for s in submissions]
        assignments = list(current_app.db.assignments.find({'_id': {'$in': assignment_ids}}))
        assignments_map = {a['_id']: a for a in assignments}

        # 3. Format the response
        history = []
        for sub in submissions:
            assignment = assignments_map.get(sub['assignment_id'])
            if assignment:
                sub['_id'] = str(sub['_id'])
                sub['student_id'] = str(sub['student_id'])
                sub['assignment_id'] = str(sub['assignment_id'])
                sub['class_id'] = str(sub['class_id'])
                sub['assignment_name'] = assignment.get('name', '未知测验')
                history.append(sub)

        return jsonify(history), 200

    except Exception as e:
        current_app.logger.error(f"获取学生测验历史时出错: {e}")
        return jsonify({'message': '获取学生测验历史时发生内部错误', 'error': str(e)}), 500
