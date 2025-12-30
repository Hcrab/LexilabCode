from flask import request, jsonify, Blueprint
from werkzeug.security import generate_password_hash
from pymongo.errors import DuplicateKeyError
from openpyxl import load_workbook
from ..extensions import users_collection, logger
from ..config import BEIJING_TZ
from .auth import admin_required
from datetime import datetime, timezone

users_bp = Blueprint('users', __name__, url_prefix='/usersdata')

@users_bp.route("", methods=['POST'])
@admin_required
def create_user(current_user):
    data = request.get_json(force=True)
    username = data.get("username")
    password = data.get("password")
    english_name = data.get("english_name")
    role = data.get("role", "user")
    if not (username and password and english_name):
        return jsonify(error="missing fields"), 400

    password_hash = generate_password_hash(password)
    user_doc = {
        "username": username,
        "password_hash": password_hash,
        "english_name": english_name,
        "role": role,
        "created_at": datetime.now(timezone.utc)
    }
    try:
        users_collection.insert_one(user_doc)
        return jsonify(ok=True), 201
    except DuplicateKeyError:
        return jsonify(error="user exists"), 409

@users_bp.route("", methods=['GET'])
@admin_required
def list_users(current_user):
    users = list(users_collection.find({}, {"password_hash": 0}).sort("username", 1))
    return jsonify(users)

@users_bp.route("/<username>", methods=['GET'])
@admin_required
def get_user(current_user, username):
    user = users_collection.find_one({"username": username}, {"password_hash": 0})
    if not user:
        return jsonify(error="not found"), 404
    return jsonify(user)

@users_bp.route("/<username>", methods=['PUT'])
@admin_required
def update_user(current_user, username):
    data = request.get_json(force=True)
    update_fields = {}
    if "password" in data and data["password"]:
        update_fields["password_hash"] = generate_password_hash(data["password"])
    if "english_name" in data:
        update_fields["english_name"] = data["english_name"]
    if "role" in data:
        update_fields["role"] = data["role"]

    if not update_fields:
        return jsonify(error="no fields to update"), 400

    result = users_collection.update_one({"username": username}, {"$set": update_fields})
    if result.matched_count == 0:
        return jsonify(error="not found"), 404
    return jsonify(ok=True)

@users_bp.route("/<username>", methods=['DELETE'])
@admin_required
def delete_user(current_user, username):
    result = users_collection.delete_one({"username": username})
    if result.deleted_count == 0:
        return jsonify(error="not found"), 404
    return jsonify(ok=True)

@users_bp.route("/batch", methods=['POST'])
@admin_required
def batch_users(current_user):
    if "file" not in request.files:
        return jsonify(error="file missing"), 400
    file = request.files["file"]
    try:
        wb = load_workbook(file, data_only=True)
        ws = wb.active
        users_to_import = []
        for r in ws.iter_rows(min_row=2, values_only=True):
            if not r or not r[0]:
                continue
            username = str(r[0]).strip()
            password = str(r[1]).strip() if r[1] else "123456"
            english_name = str(r[2]).strip() if r[2] else ""
            role = str(r[3]).strip() if len(r) > 3 and r[3] else "user"
            if username and password and english_name:
                users_to_import.append({
                    "username": username,
                    "password_hash": generate_password_hash(password),
                    "english_name": english_name,
                    "role": role
                })

        if not users_to_import:
            return jsonify(imported=0)

        count = 0
        for user_doc in users_to_import:
            try:
                users_collection.update_one(
                    {"username": user_doc["username"]},
                    {"$set": user_doc},
                    upsert=True
                )
                count += 1
            except Exception as e:
                logger.error(f"Error importing user {user_doc['username']}: {e}")
        return jsonify(imported=count)
    except Exception as e:
        logger.error(f"Failed to process batch user file: {e}")
        return jsonify(error="invalid file"), 400
